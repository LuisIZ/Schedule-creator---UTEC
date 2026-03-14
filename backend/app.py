from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import os
import io
from database import db, Course, Section, Schedule
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='../static')

# 1. SEGURIDAD: Configuración de CORS y Secret Key
# Permitimos localhost por defecto para desarrollo.
ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', 'http://localhost:3000,http://127.0.0.1:3000').split(',')
CORS(app, origins=ALLOWED_ORIGINS)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Configuration for Database
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'app.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Create Database tables on startup
with app.app_context():
    db.create_all()

@app.route('/')
def serve_index():
    return app.send_static_file('index.html')

@app.route('/api/courses', methods=['GET'])
def get_courses():
    """ 
    REST endpoint to fetch all courses from the database. 
    Can be filtered by course_type (e.g. ?type=Obligatorio) 
    """
    course_type = request.args.get('type')
    
    query = Course.query
    if course_type:
        query = query.filter_by(course_type=course_type)
        
    courses = query.all()
    return jsonify([course.to_dict() for course in courses])

# --- Color palette for courses (hex without #) ---
COURSE_COLORS = [
    '2563EB', '7C3AED', '059669', 'DC2626', 'D97706',
    'DB2777', '4F46E5', '0891B2', '65A30D', 'C026D3',
    'EA580C', '0D9488', '9333EA', 'E11D48', '2563EB',
]

def _get_course_color(index):
    """Returns a hex color for a course based on its index."""
    return COURSE_COLORS[index % len(COURSE_COLORS)]

def _time_to_row(time_str, start_hour=7):
    """Converts '08:30' to a row offset (each hour = 2 rows for 30-min granularity)."""
    h, m = map(int, time_str.split(':'))
    return (h - start_hour) * 2 + (1 if m >= 30 else 0)

@app.route('/api/export', methods=['POST'])
def export_schedule():
    """ Generates a formatted weekly schedule as an Excel file """
    data = request.json
    
    # 2. SEGURIDAD: Validación de entrada
    if not data or 'sections' not in data or not isinstance(data['sections'], list):
        return jsonify({'error': 'Formato de datos inválido'}), 400

    # Limitar a un máximo de 100 secciones y asegurar que sean enteros
    try:
        section_ids = [int(s) for s in data['sections'][:100]]
    except (ValueError, TypeError):
        return jsonify({'error': 'Los IDs de las secciones deben ser numéricos'}), 400

    if not section_ids:
         return jsonify({'error': 'No sections provided'}), 400

    # Query Database
    results = (
        db.session.query(Schedule, Section, Course)
        .select_from(Schedule)
        .join(Section, Schedule.section_id == Section.id)
        .join(Course, Section.course_code == Course.course_code)
        .filter(Section.id.in_(section_ids))
        .all()
    )

    if not results:
        return jsonify({'error': 'No schedules found for given sections'}), 404

    # ── Build the workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Horario UTEC"

    # Constants
    DAYS = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab']
    DAY_LABELS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    START_HOUR = 7
    END_HOUR = 22
    HEADER_ROW = 2        # Day headers on row 2
    TIME_START_ROW = 3    # First time slot on row 3
    TIME_COL = 1          # Column A for time labels
    DAY_START_COL = 2     # Column B for first day

    # Common styles
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    time_font = Font(name='Calibri', size=10, bold=True, color='475569')
    time_align = Alignment(horizontal='center', vertical='center')

    # ── Title Row ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=DAY_START_COL + len(DAYS) - 1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "📅 Mi Horario — UTEC"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # ── Day Headers ──
    # Time column header
    time_header = ws.cell(row=HEADER_ROW, column=TIME_COL)
    time_header.value = "Hora"
    time_header.font = header_font
    time_header.fill = header_fill
    time_header.alignment = header_align
    time_header.border = thin_border

    for i, day_label in enumerate(DAY_LABELS):
        col = DAY_START_COL + i
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.value = day_label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[HEADER_ROW].height = 28

    # ── Time Labels & Grid Lines ──
    total_slots = (END_HOUR - START_HOUR) * 2  # 30-min slots
    for slot in range(total_slots):
        row = TIME_START_ROW + slot
        hour = START_HOUR + slot // 2
        minute = '00' if slot % 2 == 0 else '30'

        # Time label in column A
        time_cell = ws.cell(row=row, column=TIME_COL)
        time_cell.value = f"{hour:02d}:{minute}"
        time_cell.font = time_font
        time_cell.alignment = time_align
        time_cell.border = thin_border

        # Empty cells with borders for the grid
        for d in range(len(DAYS)):
            col = DAY_START_COL + d
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            # Alternate row shading for readability
            if slot % 2 == 0:
                cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws.row_dimensions[row].height = 22

    # ── Place Course Blocks ──
    # Assign a color to each unique course
    unique_courses = list({course.course_code: course for _, _, course in results}.values())
    course_color_map = {c.course_code: _get_course_color(i) for i, c in enumerate(unique_courses)}

    for sched, sec, course in results:
        if not sched.day or not sched.start_time or not sched.end_time:
            continue
        if sched.day not in DAYS:
            continue

        day_idx = DAYS.index(sched.day)
        col = DAY_START_COL + day_idx

        start_slot = _time_to_row(sched.start_time, START_HOUR)
        end_slot = _time_to_row(sched.end_time, START_HOUR)
        if end_slot <= start_slot:
            end_slot = start_slot + 1  # at least 1 slot

        start_row = TIME_START_ROW + start_slot
        end_row = TIME_START_ROW + end_slot - 1

        # Merge cells for the course block
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        # Style the block
        color_hex = course_color_map[course.course_code]
        block_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        block_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
        block_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        block_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

        cell = ws.cell(row=start_row, column=col)
        cell.value = f"{course.name}\n{sec.name}\n{sched.start_time}-{sched.end_time}"
        cell.font = block_font
        cell.fill = block_fill
        cell.alignment = block_align
        cell.border = block_border

    # ── Column Widths ──
    ws.column_dimensions[get_column_letter(TIME_COL)].width = 10
    for i in range(len(DAYS)):
        ws.column_dimensions[get_column_letter(DAY_START_COL + i)].width = 22

    # ── Save to Output/ folder ──
    output_dir = os.path.join(basedir, "..", "Output")
    os.makedirs(output_dir, exist_ok=True)
    local_path = os.path.join(output_dir, "horario_armado.xlsx")
    wb.save(local_path)

    # ── Return file for download ──
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name='Mi_Horario_UTEC.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    print("Initializing Flask MVC Backend...")
    # 3. SEGURIDAD: DEBUG desactivado por defecto para evitar exposición de errores en producción
    IS_DEBUG = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=IS_DEBUG, port=5000)